from .models import Participant
from django.shortcuts import redirect
import random
from openpyxl import load_workbook

def search(request):
	q = request.GET.get('q','')
	q = q.strip()
	if q is not '':
		data = Participant.objects.filter(name__startswith = q);
	else:
		data = Participant.objects.all();

	return data,q

def storeToDB(row):
	data = [i for i in row]
	data[4] = (True if int(data[4]) is 1 else False)
	data[5] = (True if int(data[5]) is 1 else False)
	query = Participant(name=data[0],email=data[1],number=data[2],
		college=data[3],paid=data[4],present=data[5])
	query.save()

def handle_uploaded_file(f):
	file = 'uploads/'+'file_'+str(random.randrange(1000,10000))+f.name
	with open(file, 'wb+') as destination:
		for chunk in f.chunks():
			destination.write(chunk)

	"""
	The xl sheet should have only five columns inorder
	name, email, number, college, paid, present
	else this doesn't work might crash or give unusual data
	"""

	wb = load_workbook(file)
	name = wb.get_sheet_names()
	for i in name:
		sheet = wb.get_sheet_by_name(i)
		val = ((k.value for k in j) for j in sheet.iter_rows()) 
		for row in val:
			storeToDB(row)
		   



